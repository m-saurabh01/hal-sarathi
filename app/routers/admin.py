from __future__ import annotations
from fastapi import APIRouter, Depends, UploadFile, File, Form, Query
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse
from pathlib import Path
import csv
import io
from typing import List, Dict
from app.services.auth import get_admin
from app.services.data import DataService
from app.routers.public import set_data
import html as _html

router = APIRouter()


@router.get("/admin", response_class=HTMLResponse)
async def admin_ui(_: str = Depends(get_admin)):
        html = Path("app/templates/admin.html").read_text(encoding="utf-8")
        return HTMLResponse(content=html)


@router.post("/admin/upload")
async def admin_upload(
        file: UploadFile = File(...),
        mode: str = Form("replace"),
        _: str = Depends(get_admin),
):
        name = file.filename or ""
        content = await file.read()
        rows: List[Dict[str, str]] = []

        if name.lower().endswith(".csv"):
                txt = content.decode("utf-8", errors="ignore")
                reader = csv.DictReader(io.StringIO(txt))
                rows = list(reader)
        elif name.lower().endswith(".xlsx"):
                try:
                        from openpyxl import load_workbook
                except Exception:
                        return JSONResponse({"error": "XLSX support not available. Please install openpyxl."}, status_code=400)
                wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                data_iter = ws.iter_rows(values_only=True)
                try:
                        header = next(data_iter)
                except StopIteration:
                        header = []
                header = [str(h or '').strip().lower() for h in header]
                for r in data_iter:
                        row = {header[i]: ('' if v is None else str(v)) for i, v in enumerate(r) if i < len(header)}
                        rows.append(row)
        else:
                return JSONResponse({"error": "Unsupported file type. Please upload .csv or .xlsx."}, status_code=400)

        items, stats, errors = DataService.upsert_from_rows(rows, mode=mode)
        if errors:
                return JSONResponse({"errors": errors}, status_code=400)

        DataService.save_kb(items)
        set_data()  # hot-reload
        return JSONResponse({"status": "ok", "stats": stats, "mode": mode})


@router.get("/admin/unmatched", response_class=HTMLResponse)
async def admin_unmatched(raw: bool = Query(False), _: str = Depends(get_admin)):
        p = Path("data/unmatched.csv")
        if not p.exists():
                # Render page with navbar even if empty
                empty_html = """
                <!doctype html><html lang='en'>
                <head>
                        <meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
                        <title>HAL Sarathi · Unmatched</title>
                        <link href='/static/css/bootstrap.min.css' rel='stylesheet'>
                        <link href='/static/css/main.css' rel='stylesheet'>
                </head>
                <body>
                    <nav class='navbar navbar-light bg-white border-bottom'>
                        <div class='container d-flex align-items-center'>
                            <a class='navbar-brand d-flex align-items-center' href='/admin'>
                                <img src='/static/img/logo.png' class='brand-logo' width='40' height='40' alt='Logo'/>
                            </a>
                            <div class='ms-auto d-flex align-items-center gap-3 small text-muted'>
                                <span class='d-none d-md-inline'>Unmatched</span>
                                <span class='vr'></span>
                                <a href='/admin' class='text-decoration-none'>Admin</a>
                                <span class='vr'></span>
                                <a href='/' class='text-decoration-none'>User</a>
                            </div>
                        </div>
                    </nav>
                    <main class='container py-3'>
                        <div class='alert alert-secondary mb-0'>No unmatched queries yet.</div>
                    </main>
                    <script src='/static/js/bootstrap.bundle.min.js'></script>
                </body></html>
                """
                return HTMLResponse(empty_html)

        if raw:
                body = p.read_text(encoding="utf-8")
                return PlainTextResponse(content=body, media_type="text/csv; charset=utf-8")

        # Parse and render table (latest first)
        rows: List[Dict[str, str]] = []
        with p.open("r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for r in reader:
                        rows.append(r)
        rows = list(reversed(rows))
        headers = ["timestamp", "query", "top_suggestions"]

        def td(val: str) -> str:
                return _html.escape(str(val or ""))

        body_rows = "\n".join(
                f"<tr><td class='text-nowrap'>{td(r.get('timestamp'))}</td><td>{td(r.get('query'))}</td><td>{td(r.get('top_suggestions'))}</td></tr>"
                for r in rows[:500]
        )

        html = f"""
        <!doctype html><html lang='en'>
        <head>
                <meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
                <title>HAL Sarathi · Unmatched</title>
                <link href='/static/css/bootstrap.min.css' rel='stylesheet'>
                <link href='/static/css/main.css' rel='stylesheet'>
        </head>
        <body>
            <nav class='navbar navbar-light bg-white border-bottom'>
                <div class='container d-flex align-items-center'>
                    <a class='navbar-brand d-flex align-items-center' href='/admin'>
                        <img src='/static/img/logo.png' class='brand-logo' width='40' height='40' alt='Logo'/>
                    </a>
                    <div class='ms-auto d-flex align-items-center gap-3 small text-muted'>
                        <span class='d-none d-md-inline'>Unmatched</span>
                        <span class='vr'></span>
                        <a href='/admin' class='text-decoration-none'>Admin</a>
                        <span class='vr'></span>
                        <a href='/' class='text-decoration-none'>User</a>
                    </div>
                </div>
            </nav>
            <main class='container py-3'>
                <div class='d-flex align-items-center justify-content-between mb-3'>
                    <h1 class='h4 mb-0'>Unmatched Queries</h1>
                    <a class='btn btn-outline-secondary btn-sm' href='/admin/unmatched?raw=1'>Download CSV</a>
                </div>
                <div class='table-responsive'>
                    <table class='table table-sm table-striped align-middle'>
                        <thead>
                            <tr>{''.join(f"<th class='text-capitalize'>{_html.escape(h)}</th>" for h in headers)}</tr>
                        </thead>
                        <tbody>
                            {body_rows or "<tr><td colspan='3' class='text-muted'>No rows</td></tr>"}
                        </tbody>
                    </table>
                </div>
                <div class='small text-muted'>Showing {min(len(rows),500)} of {len(rows)} rows (latest first).</div>
            </main>
            <script src='/static/js/bootstrap.bundle.min.js'></script>
        </body></html>
        """
        return HTMLResponse(html)
